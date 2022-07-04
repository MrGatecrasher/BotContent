from selenium import webdriver #Ипортируем библиотеку
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
import openpyxl
import time


path = input("Введите путь к Excel файлу, пример C:\\Users\\user\\Documents\\example.xlsx  ")
wb = openpyxl.load_workbook(path, data_only=True)
# получаем активный лист
sheet1 = wb['Лист1']
wb.create_sheet("Template")
sheet2 = wb['Template']

a = str(input("Кол-во строк в первом листе: " + str(sheet1.max_row) + " | Y/N = "))
CatIn = str(input("Категория: "))

if a.upper() == "Y":
    options = webdriver.FirefoxOptions()
    options.set_headless(True)
    browser = webdriver.Firefox(options=options) #создаем объект Firefox браузера
##    browser = webdriver.Firefox()
    browser.get('URL') #посредством метода get, переходим по указаному URL
    #---Вход в акаунт---
    username = browser.find_element_by_name('username')
    password = browser.find_element_by_name('password')
    username.send_keys("Login")
    password.send_keys("Password")
    password.send_keys(Keys.RETURN)
    time.sleep(3)
    rows = sheet1.max_row + 1

    #---добавления новости---

    for i in range (1, rows):
        browser.get('addnews') #Страница редактора
        time.sleep(1)
        #---Заголовок---
        TextTitle = browser.find_element_by_name("title")
        TextTitle.send_keys(sheet1['H'+str(i)].value)
        time.sleep(0.5)
        #---Категория---
        browser.find_element_by_xpath("//span[contains(text(), 'Выберите категорию ...')]").click()
        Cat = browser.find_element_by_class_name('bs-searchbox')
        Cat.find_element_by_class_name('form-control').send_keys(CatIn + Keys.ENTER)
        time.sleep(0.5)
        #---Кнопка редактора---
        browser.find_element_by_class_name('mce-i-code').click()
        time.sleep(0.5)
        code = browser.find_element_by_class_name('mce-window')
        iframe = code.find_element_by_tag_name('iframe')
        iframe.send_keys(Keys.CONTROL + 'a')
        iframe.send_keys(Keys.DELETE)
        time.sleep(0.5)

        sheet = wb[sheet1['G'+str(i)].value]
        for num in range(sheet.max_row):
            row_num = num + 1  # нумерация строчек в экселе с 1
            sheet2['A' + str(row_num)].value = sheet['A' + str(row_num)].value  # в первый столбец "А" записываем индексы
        
        for cell in sheet2['A']:
            cell.value = cell.value.replace("index0",str(sheet1['A'+str(i)].value))
            cell.value = cell.value.replace("index1",str(sheet1['B'+str(i)].value))
            cell.value = cell.value.replace("index2",str(sheet1['C'+str(i)].value))
            cell.value = cell.value.replace("index3",str(sheet1['D'+str(i)].value))
            cell.value = cell.value.replace("index4",str(sheet1['E'+str(i)].value))
            cell.value = cell.value.replace("index5",str(sheet1['F'+str(i)].value))
            cell.value = cell.value.replace("index6",str(sheet1['A'+str(i)].value)+"x"+str(sheet1['B'+str(i)].value))
            iframe.send_keys(cell.value) #Написание кода в редактор

        
        time.sleep(1)    
        i+1
        print(i)
        code.find_element_by_tag_name("span").click()
        time.sleep(1)
##        browser.find_element_by_id("approve").click()
        browser.find_element_by_id("allow_main").click()
        browser.find_element_by_id("allow_rating").click()
        browser.find_element_by_id("allow_comm").click()
        time.sleep(0.5)
        browser.find_element_by_class_name('fa-floppy-o').click()


    browser.get('AdminPanel') #Переход на админ панель
    print('Готово!')
    browser.quit()
else:
    print("Стоп")
    quit()


